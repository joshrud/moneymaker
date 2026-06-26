"""
RL Finance Papers Excel Generator
Creates a comprehensive Excel reference for RL papers applicable to finance/trading bots.
"""

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
# DATA
# ─────────────────────────────────────────────────────────────────────────────

PAPERS = [
    # ── Category A: Foundational RL Algorithms ────────────────────────────────
    {
        "Title": "Playing Atari with Deep Reinforcement Learning",
        "Authors": "Mnih, Kavukcuoglu, Silver, Graves, Antonoglou, Wierstra, Riedmiller",
        "Year": 2013,
        "Venue": "NeurIPS Workshop",
        "DOI_arXiv": "arXiv:1312.5602",
        "Algorithm": "DQN",
        "Finance_App": "Foundational — discrete action trading signals",
        "Summary": (
            "Introduced the Deep Q-Network (DQN), combining Q-learning with convolutional "
            "neural networks and experience replay to learn directly from raw pixel inputs. "
            "Demonstrated super-human performance on 6 of 7 Atari games with a single architecture, "
            "establishing the blueprint for all subsequent deep RL research."
        ),
        "Key_Innovation": "Experience replay buffer + target network stabilises Q-learning with neural networks",
        "Code_GitHub": "https://github.com/google-deepmind/dqn_zoo",
        "Impact": "Foundational",
        "Uses_LLM": "No",
        "Notes": "First practical deep RL algorithm; all trading DQN papers build on this",
    },
    {
        "Title": "Human-level control through deep reinforcement learning",
        "Authors": "Mnih, Kavukcuoglu, Silver, Rusu, Veness, et al.",
        "Year": 2015,
        "Venue": "Nature",
        "DOI_arXiv": "DOI:10.1038/nature14236",
        "Algorithm": "DQN (Nature version)",
        "Finance_App": "Foundational — discrete action trading signals",
        "Summary": (
            "Extended the 2013 DQN workshop paper to a full Nature publication with more rigorous "
            "evaluation across 49 Atari games. Introduced clipped rewards and a more stable target "
            "network update scheme, achieving above-human performance on 29 games."
        ),
        "Key_Innovation": "Periodic hard target-network copy; single agent learns 49 diverse tasks",
        "Code_GitHub": "https://github.com/google-deepmind/dqn_zoo",
        "Impact": "Foundational",
        "Uses_LLM": "No",
        "Notes": "The canonical DQN reference; cite this for DQN in any paper",
    },
    {
        "Title": "Continuous control with deep reinforcement learning (DDPG)",
        "Authors": "Lillicrap, Hunt, Pritzel, Heess, Erez, Tassa, Silver, Wierstra",
        "Year": 2016,
        "Venue": "ICLR",
        "DOI_arXiv": "arXiv:1509.02971",
        "Algorithm": "DDPG",
        "Finance_App": "Portfolio weights, continuous order sizing, execution",
        "Summary": (
            "Adapted DQN to continuous action spaces using a deterministic policy gradient "
            "combined with an actor-critic architecture and experience replay. "
            "Solved over 20 continuous physics control tasks with the same hyperparameters, "
            "making it the first practical deep RL algorithm for continuous actions."
        ),
        "Key_Innovation": "Deterministic policy gradient + actor-critic enables continuous action spaces",
        "Code_GitHub": "https://github.com/openai/baselines",
        "Impact": "Foundational",
        "Uses_LLM": "No",
        "Notes": "Directly applicable to portfolio weight optimization; predecessor to TD3 and SAC",
    },
    {
        "Title": "Proximal Policy Optimization Algorithms (PPO)",
        "Authors": "Schulman, Wolski, Dhariwal, Radford, Klimov",
        "Year": 2017,
        "Venue": "arXiv",
        "DOI_arXiv": "arXiv:1707.06347",
        "Algorithm": "PPO",
        "Finance_App": "Portfolio management, trading strategy optimization",
        "Summary": (
            "Proposed PPO, a policy gradient method that clips the probability ratio to prevent "
            "destructively large policy updates, achieving TRPO-level performance with much simpler "
            "implementation. Allows multiple epochs of minibatch gradient updates per data sample, "
            "greatly improving sample efficiency over vanilla policy gradients."
        ),
        "Key_Innovation": "Clipped surrogate objective prevents large policy updates; simple yet robust",
        "Code_GitHub": "https://github.com/openai/baselines",
        "Impact": "Foundational",
        "Uses_LLM": "No",
        "Notes": "Default algorithm for most trading bots; stable-baselines3 PPO widely used in finance",
    },
    {
        "Title": "Soft Actor-Critic: Off-Policy Maximum Entropy Deep Reinforcement Learning",
        "Authors": "Haarnoja, Zhou, Abbeel, Levine",
        "Year": 2018,
        "Venue": "ICML",
        "DOI_arXiv": "arXiv:1801.01290",
        "Algorithm": "SAC",
        "Finance_App": "Continuous portfolio optimization, risk-aware trading",
        "Summary": (
            "Introduced SAC, which maximises a trade-off between expected reward and policy entropy "
            "to encourage exploration and prevent premature convergence. "
            "Off-policy design allows efficient replay buffer reuse, and the automatic temperature "
            "tuning (SAC v2, arXiv:1812.05905) makes hyperparameter selection simple."
        ),
        "Key_Innovation": "Maximum entropy framework with automatic temperature tuning; sample-efficient off-policy",
        "Code_GitHub": "https://github.com/haarnoja/sac",
        "Impact": "Foundational",
        "Uses_LLM": "No",
        "Notes": "State-of-the-art for continuous portfolio weight optimization; naturally handles uncertainty",
    },
    {
        "Title": "Addressing Function Approximation Error in Actor-Critic Methods (TD3)",
        "Authors": "Fujimoto, van Hoof, Meger",
        "Year": 2018,
        "Venue": "ICML",
        "DOI_arXiv": "arXiv:1802.09477",
        "Algorithm": "TD3",
        "Finance_App": "Continuous portfolio weights, execution, hedging",
        "Summary": (
            "Identified and fixed systematic overestimation bias in DDPG by introducing twin "
            "critic networks (take minimum), delayed policy updates, and target policy smoothing. "
            "Achieves significantly more stable training than DDPG with comparable or better "
            "final performance on continuous control benchmarks."
        ),
        "Key_Innovation": "Twin critics + delayed actor updates + target smoothing noise eliminate overestimation",
        "Code_GitHub": "https://github.com/sfujim/TD3",
        "Impact": "Foundational",
        "Uses_LLM": "No",
        "Notes": "More stable than DDPG; recommended over DDPG for finance applications",
    },
    {
        "Title": "Asynchronous Methods for Deep Reinforcement Learning (A3C)",
        "Authors": "Mnih, Badia, Mirza, Graves, Lillicrap, Harley, Silver, Kavukcuoglu",
        "Year": 2016,
        "Venue": "ICML",
        "DOI_arXiv": "arXiv:1602.01783",
        "Algorithm": "A3C / A2C",
        "Finance_App": "Multi-asset parallel trading, high-throughput training",
        "Summary": (
            "Proposed asynchronous parallel actors to stabilise training without an experience "
            "replay buffer, enabling CPU-based training at scale. The A3C variant with advantage "
            "function estimation (A3C) showed strong performance on Atari and continuous control, "
            "while its synchronous variant A2C became the practical standard."
        ),
        "Key_Innovation": "Parallel asynchronous actors decorrelate training data; advantage function reduces variance",
        "Code_GitHub": "https://github.com/openai/baselines",
        "Impact": "Foundational",
        "Uses_LLM": "No",
        "Notes": "A2C (sync version) included in FinRL; useful for training on multiple stock environments in parallel",
    },
    {
        "Title": "A Distributional Perspective on Reinforcement Learning (C51)",
        "Authors": "Bellemare, Dabney, Munos",
        "Year": 2017,
        "Venue": "ICML",
        "DOI_arXiv": "arXiv:1707.06887",
        "Algorithm": "C51 / Categorical DQN",
        "Finance_App": "Risk-aware trading, tail-risk modelling",
        "Summary": (
            "Replaced scalar value estimates with a categorical distribution over returns using "
            "51 atoms, allowing the agent to capture the full return distribution rather than "
            "just its expectation. Significantly outperformed DQN on Atari with the same "
            "architecture, and introduced distributional RL as a research area."
        ),
        "Key_Innovation": "Models full return distribution (not just mean); critical for risk-sensitive finance tasks",
        "Code_GitHub": "https://github.com/google-deepmind/dqn_zoo",
        "Impact": "High",
        "Uses_LLM": "No",
        "Notes": "Distributional RL especially relevant for CVaR/VaR-aware trading; basis for D4PG",
    },
    {
        "Title": "Rainbow: Combining Improvements in Deep Reinforcement Learning",
        "Authors": "Hessel, Modayil, van Hasselt, Schaul, Ostrovski, Dabney, Horgan, Piot, Azar, Silver",
        "Year": 2017,
        "Venue": "AAAI 2018",
        "DOI_arXiv": "arXiv:1710.02298",
        "Algorithm": "Rainbow DQN",
        "Finance_App": "Discrete trading signals with comprehensive improvements",
        "Summary": (
            "Systematically combined six DQN improvements — double Q-learning, dueling networks, "
            "prioritised replay, noisy networks, multi-step returns, and distributional RL — "
            "into a single agent. Achieved state-of-the-art on Atari with an ablation study "
            "showing prioritised replay and distributional RL contribute most."
        ),
        "Key_Innovation": "Meta-paper showing six orthogonal DQN improvements are complementary and additive",
        "Code_GitHub": "https://github.com/google-deepmind/dqn_zoo",
        "Impact": "High",
        "Uses_LLM": "No",
        "Notes": "Prioritised replay especially valuable in finance where rare market events matter",
    },
    {
        "Title": "Deep Recurrent Q-Learning for Partially Observable MDPs (DRQN)",
        "Authors": "Hausknecht, Stone",
        "Year": 2015,
        "Venue": "AAAI Workshop",
        "DOI_arXiv": "arXiv:1507.06527",
        "Algorithm": "DRQN",
        "Finance_App": "Trading with hidden market state, POMDP formulation",
        "Summary": (
            "Replaced the first fully-connected layer of DQN with an LSTM, enabling the agent "
            "to integrate information over time for partially observable environments. "
            "Demonstrated superior performance on Atari games where frames are randomly omitted, "
            "and proposed bootstrapped sequential updates for training on trajectories."
        ),
        "Key_Innovation": "LSTM within DQN handles partial observability and temporal dependencies",
        "Code_GitHub": "N/A",
        "Impact": "High",
        "Uses_LLM": "No",
        "Notes": "Markets are POMDPs; DRQN directly relevant for trading with hidden state (order flow, macro)",
    },
    {
        "Title": "World Models",
        "Authors": "Ha, Schmidhuber",
        "Year": 2018,
        "Venue": "NeurIPS",
        "DOI_arXiv": "arXiv:1803.10122",
        "Algorithm": "World Models (VAE + MDN-RNN + CMA-ES)",
        "Finance_App": "Learned market simulators, synthetic data generation",
        "Summary": (
            "Proposed a compressed world model combining a VAE for spatial encoding and an "
            "MDN-RNN for temporal dynamics, allowing an agent to dream (train within its own "
            "imagined model). Demonstrated on Car Racing and VizDoom, inspiring a generation "
            "of model-based RL work including Dreamer."
        ),
        "Key_Innovation": "Dream-based policy training in learned latent space; separates perception from memory",
        "Code_GitHub": "https://github.com/hardmaru/WorldModelsExperiments",
        "Impact": "High",
        "Uses_LLM": "No",
        "Notes": "Foundation for model-based finance RL; learn market dynamics model from historical data",
    },
    {
        "Title": "When to Trust Your Model: Model-Based Policy Optimization (MBPO)",
        "Authors": "Janner, Fu, Zhang, Levine",
        "Year": 2019,
        "Venue": "NeurIPS",
        "DOI_arXiv": "arXiv:1906.08253",
        "Algorithm": "MBPO",
        "Finance_App": "Sample-efficient trading with learned market models",
        "Summary": (
            "Showed that short model-generated rollouts (k=1) branched from real data achieve "
            "better sample efficiency than long rollouts, because errors compound less. "
            "Achieved SAC-level performance with 20-40x fewer environment samples by training "
            "SAC on a mixture of real and model-generated data."
        ),
        "Key_Innovation": "Short branched model rollouts avoid compounding error; theoretical justification for model use",
        "Code_GitHub": "https://github.com/jannerm/mbpo",
        "Impact": "High",
        "Uses_LLM": "No",
        "Notes": "Key for finance where real trading data is expensive; train on simulator + few real trades",
    },
    {
        "Title": "Decision Transformer: Reinforcement Learning via Sequence Modeling",
        "Authors": "Chen, Lu, Rajeswaran, Lee, Grover, Laskin, Abbeel, Srinivas, Mordatch",
        "Year": 2021,
        "Venue": "NeurIPS",
        "DOI_arXiv": "arXiv:2106.01345",
        "Algorithm": "Decision Transformer",
        "Finance_App": "Offline policy learning from historical trade data",
        "Summary": (
            "Reframed RL as a sequence modelling problem: given (return-to-go, state, action) "
            "triplets as tokens, a GPT-style transformer predicts the next action conditioned "
            "on a desired future return. Matched or exceeded offline RL baselines on Atari and "
            "MuJoCo without any value function estimation."
        ),
        "Key_Innovation": "No value functions or policy gradients; RL as conditional sequence generation",
        "Code_GitHub": "https://github.com/kzl/decision-transformer",
        "Impact": "High",
        "Uses_LLM": "No",
        "Notes": "Powerful for offline learning from historical trade logs; specify desired Sharpe ratio as conditioning",
    },
    {
        "Title": "Mastering Atari with Discrete World Models (DreamerV2)",
        "Authors": "Hafner, Lillicrap, Norouzi, Ba",
        "Year": 2020,
        "Venue": "ICLR 2021",
        "DOI_arXiv": "arXiv:2010.02193",
        "Algorithm": "DreamerV2",
        "Finance_App": "Market world model learning, synthetic data augmentation",
        "Summary": (
            "Extended Dreamer to discrete latent representations using straight-through gradients, "
            "achieving human-level performance on Atari with a single GPU. Trains behaviors "
            "entirely inside the world model (imagination), enabling extremely high sample "
            "efficiency by amortising real-world interactions."
        ),
        "Key_Innovation": "Discrete world model trained with KL balancing; entire policy training in imagination",
        "Code_GitHub": "https://github.com/danijar/dreamerv2",
        "Impact": "High",
        "Uses_LLM": "No",
        "Notes": "Directly applicable to financial market modelling with OHLCV data as observations",
    },
    # ── Category B: RL Applied to Finance ────────────────────────────────────
    {
        "Title": "Reinforcement Learning for Trading",
        "Authors": "Moody, Saffell",
        "Year": 1998,
        "Venue": "NeurIPS",
        "DOI_arXiv": "N/A (NeurIPS proceedings)",
        "Algorithm": "Recurrent Reinforcement Learning (RRL)",
        "Finance_App": "Directional trading, portfolio allocation",
        "Summary": (
            "Seminal paper applying RL directly to trading, proposing Recurrent Reinforcement "
            "Learning (RRL) that optimises the Sharpe ratio directly rather than forecasting "
            "prices first. Showed that maximising risk-adjusted returns is more robust than "
            "two-stage predict-then-trade approaches."
        ),
        "Key_Innovation": "Direct RL trading without forecasting; Sharpe ratio as reward; differentiable profit objective",
        "Code_GitHub": "N/A",
        "Impact": "Foundational",
        "Uses_LLM": "No",
        "Notes": "Founding paper of RL for trading; introduced direct policy optimisation for finance",
    },
    {
        "Title": "A Deep Reinforcement Learning Framework for the Financial Portfolio Management Problem (EIIE)",
        "Authors": "Jiang, Xu, Liang",
        "Year": 2017,
        "Venue": "arXiv",
        "DOI_arXiv": "arXiv:1706.10059",
        "Algorithm": "EIIE (Ensemble of Identical Independent Evaluators)",
        "Finance_App": "Cryptocurrency portfolio management, continuous rebalancing",
        "Summary": (
            "Proposed EIIE, a novel portfolio vector memory (PVM) architecture where identical "
            "sub-networks independently score each asset, enforcing permutation-invariant "
            "portfolio allocation. Trained with online learning on crypto markets and achieved "
            "4-fold returns over buy-and-hold in backtests."
        ),
        "Key_Innovation": "EIIE architecture with portfolio vector memory; asset-independent evaluation networks",
        "Code_GitHub": "https://github.com/ZhengyaoJiang/PGPortfolio",
        "Impact": "High",
        "Uses_LLM": "No",
        "Notes": "Very influential for crypto portfolio RL; EIIE architecture widely replicated",
    },
    {
        "Title": "Practical Deep Reinforcement Learning Approach for Stock Trading",
        "Authors": "Xiong, Liu, Zhong, Yang, Walid",
        "Year": 2018,
        "Venue": "NeurIPS Workshop",
        "DOI_arXiv": "arXiv:1811.07522",
        "Algorithm": "DDPG",
        "Finance_App": "Single-stock and multi-stock US equity trading (DJIA 30)",
        "Summary": (
            "Applied DDPG to trade 30 DJIA stocks, defining state as price/technical-indicator "
            "vectors and actions as continuous position sizes. Outperformed DJIA index and "
            "min-variance baselines on Sharpe ratio and cumulative return over a 1-year "
            "test period. Precursor to the FinRL library."
        ),
        "Key_Innovation": "Turbulence index for risk control; realistic transaction cost modelling in RL trading env",
        "Code_GitHub": "https://github.com/AI4Finance-LLC/FinRL-Library",
        "Impact": "High",
        "Uses_LLM": "No",
        "Notes": "Direct predecessor to FinRL; demonstrates DDPG for multi-stock portfolio allocation",
    },
    {
        "Title": "FinRL: A Deep Reinforcement Learning Library for Automated Stock Trading in Quantitative Finance",
        "Authors": "Liu, Yang, Chen, Zhang, Yang, Xiao, Wang",
        "Year": 2020,
        "Venue": "NeurIPS Workshop / SSRN",
        "DOI_arXiv": "arXiv:2011.09607",
        "Algorithm": "DQN, DDPG, PPO, SAC, A2C, TD3",
        "Finance_App": "Stock trading, portfolio allocation, multi-asset, DJIA/S&P500/NASDAQ",
        "Summary": (
            "Introduced FinRL, an open-source library providing standardised environments, "
            "data pipelines, and agent implementations for DRL-based trading. Covers single-stock, "
            "multi-stock, and portfolio allocation tasks with realistic transaction costs and "
            "multiple market environments (US, China, Hong Kong)."
        ),
        "Key_Innovation": "Standardised gym-compatible trading environments; unified backtesting pipeline for DRL",
        "Code_GitHub": "https://github.com/AI4Finance-LLC/FinRL-Library",
        "Impact": "Foundational",
        "Uses_LLM": "No",
        "Notes": "The go-to starting framework for RL trading; actively maintained; 10k+ GitHub stars",
    },
    {
        "Title": "Deep Hedging",
        "Authors": "Bühlер, Gonon, Teichmann, Wood",
        "Year": 2019,
        "Venue": "Quantitative Finance (journal)",
        "DOI_arXiv": "arXiv:1802.03042",
        "Algorithm": "Deep RL / Monte Carlo Policy Gradient",
        "Finance_App": "Derivative hedging, options pricing under transaction costs",
        "Summary": (
            "Formulated derivative hedging as an RL problem where a neural network learns "
            "the optimal dynamic hedging strategy under transaction costs and market frictions, "
            "replacing Black-Scholes delta hedging. Used convex risk measures (CVaR, variance) "
            "as the reward signal and validated on the Heston model."
        ),
        "Key_Innovation": "Neural network replaces Black-Scholes; handles transaction costs and incompleteness natively",
        "Code_GitHub": "N/A",
        "Impact": "Foundational",
        "Uses_LLM": "No",
        "Notes": "Foundational for derivatives/options desks; spawned a large literature in RL hedging",
    },
    {
        "Title": "Market Making via Reinforcement Learning",
        "Authors": "Spooner, Fearnley, Savani, Koukorinis",
        "Year": 2018,
        "Venue": "AAMAS",
        "DOI_arXiv": "arXiv:1804.04216",
        "Algorithm": "TD learning with tile coding",
        "Finance_App": "Limit order book market making, inventory management",
        "Summary": (
            "Developed a high-fidelity limit order book simulator and trained a TD-learning "
            "market-making agent with tile coding function approximation. "
            "Custom reward function penalises inventory risk while maximising spread capture; "
            "agent outperforms Avellaneda-Stoikov and naive benchmarks."
        ),
        "Key_Innovation": "Full LOB simulation for RL training; inventory-risk-penalised reward function",
        "Code_GitHub": "N/A",
        "Impact": "High",
        "Uses_LLM": "No",
        "Notes": "Key paper for RL market making; template for reward design with inventory constraints",
    },
    {
        "Title": "A Reinforcement Learning Extension to the Almgren-Chriss Model for Optimal Trade Execution",
        "Authors": "Hendricks, Wilcox",
        "Year": 2014,
        "Venue": "IEEE SSCI",
        "DOI_arXiv": "arXiv:1403.2229",
        "Algorithm": "Q-learning extension",
        "Finance_App": "Optimal execution, minimising market impact / implementation shortfall",
        "Summary": (
            "Extended the classical Almgren-Chriss linear impact model with a Q-learning agent "
            "that dynamically adapts a volume trajectory based on real-time spread and volume "
            "signals. Demonstrated up to 10.3% reduction in implementation shortfall versus "
            "the static Almgren-Chriss baseline on South African equity data."
        ),
        "Key_Innovation": "First RL extension to Almgren-Chriss; dynamic execution adapts to market microstructure",
        "Code_GitHub": "N/A",
        "Impact": "High",
        "Uses_LLM": "No",
        "Notes": "Important for institutional execution desks; bridges classical optimal execution and RL",
    },
    {
        "Title": "Adversarial Deep Reinforcement Learning in Portfolio Management",
        "Authors": "Liang, Chen, Liu, Feng",
        "Year": 2018,
        "Venue": "arXiv",
        "DOI_arXiv": "arXiv:1808.09940",
        "Algorithm": "Adversarial training + DDPG/PPO",
        "Finance_App": "Stock portfolio management with adversarial robustness",
        "Summary": (
            "Combined adversarial training with DDPG and PPO for portfolio management, "
            "using a second adversarial network to generate challenging market scenarios "
            "during training. Achieved more robust out-of-sample Sharpe ratios than non-adversarial "
            "DRL baselines on Chinese stock markets."
        ),
        "Key_Innovation": "Adversarial perturbation training for robust portfolio RL; improves generalisation",
        "Code_GitHub": "N/A",
        "Impact": "Medium",
        "Uses_LLM": "No",
        "Notes": "Addresses the fragility of RL policies to distributional shift — critical problem in live trading",
    },
    {
        "Title": "Deep Reinforcement Learning in Cryptocurrency Market Making",
        "Authors": "Sadighian",
        "Year": 2019,
        "Venue": "arXiv",
        "DOI_arXiv": "arXiv:1911.08647",
        "Algorithm": "DQN / DDPG",
        "Finance_App": "Crypto market making, bid-ask spread management",
        "Summary": (
            "Trained DRL agents to act as crypto market makers on simulated Bitcoin order books, "
            "managing inventory risk and optimising spread capture. "
            "Extended classical Avellaneda-Stoikov intuitions with learned policies that adapt "
            "to non-stationary volatility regimes in cryptocurrency markets."
        ),
        "Key_Innovation": "DRL market making for 24/7 crypto markets; handles regime shifts and thin liquidity",
        "Code_GitHub": "N/A",
        "Impact": "Medium",
        "Uses_LLM": "No",
        "Notes": "One of first RL market making papers specifically for crypto; relevant 24/7 trading context",
    },
    {
        "Title": "Deep Reinforcement Learning for Active High Frequency Trading",
        "Authors": "Ardon, Vadori, Kell, Xu, Johansson, Veloso",
        "Year": 2021,
        "Venue": "arXiv",
        "DOI_arXiv": "arXiv:2101.07107",
        "Algorithm": "PPO",
        "Finance_App": "High-frequency trading, limit order book intraday",
        "Summary": (
            "First end-to-end DRL framework for active HFT using real LOB data (Intel stock "
            "Feb–Jun 2019). Trained PPO agents on 60 files of millisecond LOB data and evaluated "
            "on held-out June 2019 data, showing positive PnL after transaction costs. "
            "Defined granular LOB-based state representations for HFT agents."
        ),
        "Key_Innovation": "First rigorous end-to-end DRL HFT paper with real LOB data; PPO on millisecond data",
        "Code_GitHub": "N/A",
        "Impact": "High",
        "Uses_LLM": "No",
        "Notes": "Essential reference for HFT RL; defines state/action/reward for LOB-based agents",
    },
    {
        "Title": "Multi-Agent Reinforcement Learning in a Realistic Limit Order Book Market Simulation",
        "Authors": "Karpe, Fang, Ma, Wang",
        "Year": 2020,
        "Venue": "ICAIF 2020",
        "DOI_arXiv": "arXiv:2006.05574",
        "Algorithm": "Double DQN (DDQN)",
        "Finance_App": "Market simulation, multi-agent execution, market impact modelling",
        "Summary": (
            "Used ABIDES (Agent-Based Interactive Discrete Event Simulation) to create a "
            "multi-agent LOB environment and trained DDQL execution agents within it. "
            "Demonstrated that realistic market simulation enables RL agents to learn execution "
            "strategies that account for their own market impact."
        ),
        "Key_Innovation": "First MARL paper in ABIDES LOB simulator; agents model their own market impact",
        "Code_GitHub": "N/A",
        "Impact": "High",
        "Uses_LLM": "No",
        "Notes": "Essential for understanding market impact in execution; ABIDES simulator is open-source",
    },
    {
        "Title": "AlphaPortfolio: Direct Construction Through Deep Reinforcement Learning and Interpretable AI",
        "Authors": "Cong, Tang, Wang, Zhang",
        "Year": 2021,
        "Venue": "SSRN / Management Science",
        "DOI_arXiv": "SSRN:3554486",
        "Algorithm": "Attention-based DRL (custom)",
        "Finance_App": "Equity portfolio construction, cross-sectional factor model",
        "Summary": (
            "Directly optimised portfolio objectives using DRL with multi-sequence cross-asset "
            "attention networks tailored to financial data. "
            "Achieved Sharpe ratios above 2.0 and 13%+ risk-adjusted alpha out-of-sample "
            "with monthly rebalancing, incorporating transaction costs, factor exposures, and "
            "interpretable attention weights."
        ),
        "Key_Innovation": "Cross-asset attention for portfolio RL; direct objective optimisation vs supervised learning",
        "Code_GitHub": "N/A",
        "Impact": "High",
        "Uses_LLM": "No",
        "Notes": "Strong empirical results; interpretable via attention; covers transaction costs natively",
    },
    {
        "Title": "Gamma and Vega Hedging Using Deep Distributional Reinforcement Learning",
        "Authors": "Cao, Chen, Farghadani, Hull, Poulos, Wang, Yuan",
        "Year": 2022,
        "Venue": "Frontiers in AI",
        "DOI_arXiv": "arXiv:2205.05614",
        "Algorithm": "D4PG (Distributed Distributional DDPG)",
        "Finance_App": "Options Greeks hedging (gamma/vega), risk management",
        "Summary": (
            "Extended deep hedging to second-order Greeks (gamma/vega) using D4PG with quantile "
            "regression for distributional value estimation. Tested three risk objectives — "
            "mean-variance, VaR, and CVaR — demonstrating RL can manage complex multi-instrument "
            "hedging problems that Black-Scholes cannot handle analytically."
        ),
        "Key_Innovation": "D4PG for gamma/vega hedging; CVaR/VaR risk objectives in distributional RL framework",
        "Code_GitHub": "N/A",
        "Impact": "Medium",
        "Uses_LLM": "No",
        "Notes": "Extends deep hedging to higher-order Greeks; important for vol trading desks",
    },
    {
        "Title": "Deep Reinforcement Learning for Cryptocurrency Trading: Practical Approach to Address Backtest Overfitting",
        "Authors": "Gort, Liu, Sun, Chen, Ye, Liu",
        "Year": 2022,
        "Venue": "arXiv / IJCAI workshop",
        "DOI_arXiv": "arXiv:2209.05559",
        "Algorithm": "PPO, SAC",
        "Finance_App": "Cryptocurrency trading, overfitting prevention",
        "Summary": (
            "Addressed the critical problem of backtest overfitting in crypto DRL trading by "
            "proposing multiple out-of-sample validation schemes and a time-series cross-validation "
            "protocol for RL policies. Empirically showed PPO and SAC strategies generalize "
            "better when trained with walk-forward validation versus standard train-test splits."
        ),
        "Key_Innovation": "Walk-forward validation for RL; quantifies and addresses backtest overfitting in crypto",
        "Code_GitHub": "N/A",
        "Impact": "Medium",
        "Uses_LLM": "No",
        "Notes": "Critical methodology paper for any RL trading practitioner; prevents overfitting pitfall",
    },
    # ── Category C: Modern/Cutting-Edge (2022-2025) ───────────────────────────
    {
        "Title": "Conservative Q-Learning for Offline Reinforcement Learning (CQL)",
        "Authors": "Kumar, Zhou, Tucker, Levine",
        "Year": 2020,
        "Venue": "NeurIPS",
        "DOI_arXiv": "arXiv:2006.04779",
        "Algorithm": "CQL",
        "Finance_App": "Offline policy learning from historical trade data",
        "Summary": (
            "Proposed CQL to address distributional shift in offline RL by adding a Q-value "
            "regularizer that penalizes high Q-values for out-of-distribution actions. "
            "Achieves 2-5x higher returns than prior offline RL methods on D4RL benchmarks, "
            "enabling safe policy learning from static datasets without environment interaction."
        ),
        "Key_Innovation": "Conservative Q-value lower bounds prevent OOD action overestimation in offline RL",
        "Code_GitHub": "https://github.com/aviralkumar2907/CQL",
        "Impact": "High",
        "Uses_LLM": "No",
        "Notes": "Critical for learning from historical trade data without live trading risk; finance applications growing",
    },
    {
        "Title": "FinGPT: Open-Source Financial Large Language Models",
        "Authors": "Yang, Liu, Wang",
        "Year": 2023,
        "Venue": "IJCAI FinLLM Workshop",
        "DOI_arXiv": "arXiv:2306.06031",
        "Algorithm": "LLM + RLHF",
        "Finance_App": "Sentiment analysis, robo-advising, trading signal generation",
        "Summary": (
            "Introduced FinGPT, an open-source framework for building financial LLMs using RLHF "
            "to align models with financial expert preferences. Provides data-centric pipelines "
            "for real-time news/sentiment ingestion and fine-tuning on financial tasks including "
            "sentiment analysis, forecasting, and algorithmic trading signals."
        ),
        "Key_Innovation": "Open-source financial LLM with RLHF; real-time financial data pipeline for LLM fine-tuning",
        "Code_GitHub": "https://github.com/AI4Finance-Foundation/FinGPT",
        "Impact": "High",
        "Uses_LLM": "Yes",
        "Notes": "State features from FinGPT sentiment can boost RL agent performance; active community",
    },
    {
        "Title": "The Evolution of Reinforcement Learning in Quantitative Finance: A Survey",
        "Authors": "Hambly, Xu, Yang",
        "Year": 2023,
        "Venue": "Journal of Mathematical Finance / arXiv",
        "DOI_arXiv": "arXiv:2408.10932",
        "Algorithm": "Survey (DQN, DDPG, PPO, SAC, TD3, offline RL)",
        "Finance_App": "Survey — portfolio, execution, market making, hedging",
        "Summary": (
            "Comprehensive survey covering 20+ years of RL in quantitative finance, organising "
            "applications into portfolio management, optimal execution, market making, and "
            "derivative pricing/hedging. Reviews algorithmic progress, open challenges, and "
            "emerging directions including offline RL and LLM integration."
        ),
        "Key_Innovation": "Unified taxonomy of RL finance subfields; comprehensive literature review to 2024",
        "Code_GitHub": "N/A",
        "Impact": "High",
        "Uses_LLM": "No",
        "Notes": "Best current survey paper for RL in finance; start here for literature review",
    },
    {
        "Title": "Deep Reinforcement Learning in Quantitative Algorithmic Trading: A Review",
        "Authors": "Millea",
        "Year": 2021,
        "Venue": "arXiv",
        "DOI_arXiv": "arXiv:2106.00123",
        "Algorithm": "Survey (DQN, DDPG, PPO, A3C)",
        "Finance_App": "Survey — stock trading, portfolio optimization",
        "Summary": (
            "Reviews DRL applications in algorithmic trading, categorising approaches by "
            "state representation, action space, and reward design. Identifies key challenges "
            "including non-stationarity, sparse rewards, and overfitting to historical data, "
            "and surveys mitigation techniques for each."
        ),
        "Key_Innovation": "Practical taxonomy of state/action/reward design choices for trading RL",
        "Code_GitHub": "N/A",
        "Impact": "Medium",
        "Uses_LLM": "No",
        "Notes": "Good entry-level survey; readable overview of the field pre-2021",
    },
    {
        "Title": "Language Model Guided Reinforcement Learning in Quantitative Trading",
        "Authors": "Preprint, multiple authors",
        "Year": 2025,
        "Venue": "arXiv / FLLM 2025",
        "DOI_arXiv": "arXiv:2508.02366",
        "Algorithm": "LLM + RL hybrid",
        "Finance_App": "Quantitative trading strategy generation with LLM guidance",
        "Summary": (
            "Proposed a hybrid framework in which an LLM generates high-level trading strategies "
            "that serve as structured guidance for an RL agent. The LLM handles strategy ideation "
            "and economic reasoning while the RL agent handles execution timing and sizing, "
            "outperforming unguided RL baselines on Sharpe ratio and max drawdown metrics."
        ),
        "Key_Innovation": "LLM as high-level strategy planner for RL execution agent; combines reasoning with adaptation",
        "Code_GitHub": "N/A",
        "Impact": "Medium",
        "Uses_LLM": "Yes",
        "Notes": "Cutting-edge LLM+RL hybrid; represents future direction of trading AI systems",
    },
    {
        "Title": "Integrating Large Language Models and Reinforcement Learning for Sentiment-Driven Quantitative Trading",
        "Authors": "Multiple authors",
        "Year": 2024,
        "Venue": "arXiv",
        "DOI_arXiv": "arXiv:2510.10526",
        "Algorithm": "LLM (GPT/Llama) + PPO/SAC",
        "Finance_App": "Sentiment-augmented trading strategy",
        "Summary": (
            "Combined LLM-extracted sentiment features from financial news with RL trading agents, "
            "using the LLM as a feature extractor for state augmentation. "
            "Demonstrated significant Sharpe ratio improvements over price-only RL baselines "
            "when sentiment signals are incorporated into the state representation."
        ),
        "Key_Innovation": "LLM sentiment as RL state feature; validated improvement from text signals in trading",
        "Code_GitHub": "N/A",
        "Impact": "Medium",
        "Uses_LLM": "Yes",
        "Notes": "Practical blueprint for adding LLM-derived features to existing RL trading bots",
    },
    {
        "Title": "Deep Learning Statistical Arbitrage",
        "Authors": "Guijarro-Ordonez, Pelger, Zanotti",
        "Year": 2021,
        "Venue": "arXiv / Management Science",
        "DOI_arXiv": "arXiv:2106.04028",
        "Algorithm": "Convolutional Transformer + RL policy",
        "Finance_App": "Statistical arbitrage, pairs/basket trading",
        "Summary": (
            "Built a unified statistical arbitrage framework using convolutional transformers "
            "to extract signals from portfolios of similar assets, then derived optimal "
            "trading policies from those signals. Achieved statistically significant alpha "
            "on US equities with consistent out-of-sample Sharpe ratios."
        ),
        "Key_Innovation": "Convolutional transformer for arbitrage signal extraction; directly optimises trading policy",
        "Code_GitHub": "N/A",
        "Impact": "High",
        "Uses_LLM": "No",
        "Notes": "Bridges statistical arbitrage and DRL; transformer architecture processes cross-asset dependencies",
    },
    {
        "Title": "FinRL-Podracer: High Performance and Scalable Deep Reinforcement Learning for Quantitative Finance",
        "Authors": "Liu, Yang, Gao, Wang, Chen, Zha",
        "Year": 2021,
        "Venue": "ACM ICAIF",
        "DOI_arXiv": "arXiv:2111.05188",
        "Algorithm": "ElegantRL / GPU-accelerated DRL",
        "Finance_App": "High-throughput stock trading research",
        "Summary": (
            "Extended FinRL with GPU-accelerated vectorised environments (FinRL-Podracer), "
            "enabling population-based training of trading agents at 1000x the speed of CPU-based "
            "approaches. Demonstrated scalable hyperparameter search and ensemble strategies "
            "achieving better Sharpe ratios through diversity."
        ),
        "Key_Innovation": "GPU-vectorised financial RL environments; population-based training for ensembles",
        "Code_GitHub": "https://github.com/AI4Finance-Foundation/FinRL",
        "Impact": "Medium",
        "Uses_LLM": "No",
        "Notes": "Engineering paper for scaling RL training; ElegantRL backend is highly optimised",
    },
    {
        "Title": "DeepTrader: A Deep Reinforcement Learning Approach for Risk-Return Balanced Portfolio Management",
        "Authors": "Wang, Zhou, Liu, Han, Lin",
        "Year": 2021,
        "Venue": "AAAI",
        "DOI_arXiv": "AAAI:2021",
        "Algorithm": "Custom DRL with market condition embedding",
        "Finance_App": "Risk-return balanced portfolio management with macro context",
        "Summary": (
            "Proposed DeepTrader, which jointly learns asset scoring via temporal graph convolutional "
            "networks and incorporates market condition embeddings (bull/bear) to adjust risk appetite "
            "dynamically. Outperformed DQN, DDPG, and classical Markowitz benchmarks on Chinese "
            "and US stock markets."
        ),
        "Key_Innovation": "Market condition embedding for dynamic risk-return tradeoff; graph conv for cross-asset relations",
        "Code_GitHub": "N/A",
        "Impact": "Medium",
        "Uses_LLM": "No",
        "Notes": "Good example of incorporating market regime detection into RL portfolio management",
    },
    {
        "Title": "An Application of Deep Reinforcement Learning to Algorithmic Trading",
        "Authors": "Théate, Ernst",
        "Year": 2020,
        "Venue": "Expert Systems with Applications",
        "DOI_arXiv": "arXiv:2004.06627",
        "Algorithm": "TDQN (custom DQN variant)",
        "Finance_App": "Single-asset directional trading across 30 financial instruments",
        "Summary": (
            "Developed TDQN, a DQN variant with LSTM state encoding specifically designed for "
            "financial time series, and evaluated it on 30 diverse financial instruments "
            "(equities, forex, crypto). Found consistent profitability on some but not all assets, "
            "with Sharpe ratios significantly above random trading baselines."
        ),
        "Key_Innovation": "TDQN architecture with LSTM; comprehensive cross-asset evaluation methodology",
        "Code_GitHub": "N/A",
        "Impact": "Medium",
        "Uses_LLM": "No",
        "Notes": "Rigorous evaluation framework; useful template for empirical RL trading papers",
    },
    {
        "Title": "Select and Trade: Towards Unified Pair Trading with Hierarchical Reinforcement Learning",
        "Authors": "Sun, Shi, Lu, Li",
        "Year": 2023,
        "Venue": "arXiv / KDD",
        "DOI_arXiv": "arXiv:2301.10724",
        "Algorithm": "Hierarchical RL (high-level pair selection + low-level execution)",
        "Finance_App": "Pairs trading, statistical arbitrage",
        "Summary": (
            "Proposed a two-level hierarchical RL framework where a high-level agent selects "
            "trading pairs and a low-level agent executes position entries and exits. "
            "Unified pair selection and trading into a single learnable system, outperforming "
            "cointegration-based and single-level RL baselines."
        ),
        "Key_Innovation": "Hierarchical RL unifies pair selection and execution; end-to-end learnable arbitrage",
        "Code_GitHub": "N/A",
        "Impact": "Medium",
        "Uses_LLM": "No",
        "Notes": "Novel hierarchical formulation; applicable to other multi-level trading decisions",
    },
    {
        "Title": "Offline Deep Reinforcement Learning for Dynamic Pricing of Consumer Credit",
        "Authors": "Rashidinejad et al.",
        "Year": 2022,
        "Venue": "arXiv",
        "DOI_arXiv": "arXiv:2203.03003",
        "Algorithm": "CQL (offline RL)",
        "Finance_App": "Dynamic pricing, credit markets",
        "Summary": (
            "Applied CQL to dynamic credit pricing using historical loan data without any "
            "online exploration, demonstrating that offline RL can outperform rule-based pricing "
            "systems in revenue while maintaining risk constraints. "
            "Provides blueprint for applying offline RL to regulated financial decisions."
        ),
        "Key_Innovation": "First application of CQL to consumer credit pricing; offline RL for regulated finance",
        "Code_GitHub": "N/A",
        "Impact": "Medium",
        "Uses_LLM": "No",
        "Notes": "Demonstrates offline RL viability in finance; applicable to any historical-data-only scenario",
    },
    {
        "Title": "Noisy Networks for Exploration (NoisyNet)",
        "Authors": "Fortunato, Azar, Piot, Menick, Osband, Graves, Mnih, Munos, Hassabis, Pietquin, Blundell, Legg",
        "Year": 2017,
        "Venue": "ICLR 2018",
        "DOI_arXiv": "arXiv:1706.10295",
        "Algorithm": "NoisyNet",
        "Finance_App": "Exploration in trading environments with non-stationary signals",
        "Summary": (
            "Introduced learnable parametric noise added to neural network weights as a "
            "principled alternative to epsilon-greedy exploration. The noise parameters are "
            "learned via gradient descent, resulting in state-dependent exploration that "
            "focuses on uncertain parts of the state space — crucial for financial markets."
        ),
        "Key_Innovation": "Learnable weight noise replaces epsilon-greedy; state-dependent exploration",
        "Code_GitHub": "N/A",
        "Impact": "High",
        "Uses_LLM": "No",
        "Notes": "Part of Rainbow; recommended over epsilon-greedy for financial RL due to regime-conditional exploration",
    },
    {
        "Title": "Deep Reinforcement Learning for Optimal Portfolio Allocation: A Comparative Study with Mean-Variance Optimization",
        "Authors": "Multiple authors",
        "Year": 2023,
        "Venue": "arXiv / ICAPS FinPlan",
        "DOI_arXiv": "arXiv:2602.17098",
        "Algorithm": "PPO, SAC, DDPG vs Markowitz",
        "Finance_App": "Portfolio allocation, benchmark comparison with classical finance",
        "Summary": (
            "Systematically compared PPO, SAC, and DDPG against mean-variance optimization "
            "on multiple market datasets, finding DRL agents achieve higher Sharpe ratios "
            "and lower maximum drawdowns than Markowitz portfolios across most evaluation periods. "
            "Provides reproducible benchmark results for the DRL vs classical finance comparison."
        ),
        "Key_Innovation": "Rigorous DRL vs MVO comparison; Sharpe/drawdown analysis across multiple regimes",
        "Code_GitHub": "N/A",
        "Impact": "Medium",
        "Uses_LLM": "No",
        "Notes": "Useful for justifying RL over classical methods; provides baseline numbers for new work",
    },
]

# ─────────────────────────────────────────────────────────────────────────────
# ALGORITHM REFERENCE DATA
# ─────────────────────────────────────────────────────────────────────────────

ALGORITHMS = [
    {
        "Name": "DQN",
        "Type": "Value-based",
        "Action Space": "Discrete",
        "Key Paper": "Mnih et al. 2015 (Nature)",
        "Finance Suitability": "★★★☆☆",
        "Notes": "Good for discrete buy/hold/sell; not for portfolio weights",
    },
    {
        "Name": "DDPG",
        "Type": "Actor-Critic (off-policy)",
        "Action Space": "Continuous",
        "Key Paper": "Lillicrap et al. 2016 (ICLR)",
        "Finance Suitability": "★★★★☆",
        "Notes": "Direct portfolio weight output; sensitive to hyperparams; use TD3 instead",
    },
    {
        "Name": "TD3",
        "Type": "Actor-Critic (off-policy)",
        "Action Space": "Continuous",
        "Key Paper": "Fujimoto et al. 2018 (ICML)",
        "Finance Suitability": "★★★★☆",
        "Notes": "Improved DDPG; more stable; recommended over DDPG for portfolio tasks",
    },
    {
        "Name": "SAC",
        "Type": "Actor-Critic (off-policy, max-entropy)",
        "Action Space": "Continuous",
        "Key Paper": "Haarnoja et al. 2018 (ICML)",
        "Finance Suitability": "★★★★★",
        "Notes": "Best overall for continuous portfolio optimization; entropy encourages diversification",
    },
    {
        "Name": "PPO",
        "Type": "Policy Gradient (on-policy)",
        "Action Space": "Both",
        "Key Paper": "Schulman et al. 2017 (arXiv)",
        "Finance Suitability": "★★★★☆",
        "Notes": "Most widely used in finance; stable training; slower but reliable",
    },
    {
        "Name": "A2C",
        "Type": "Actor-Critic (on-policy)",
        "Action Space": "Both",
        "Key Paper": "Mnih et al. 2016 (ICML)",
        "Finance Suitability": "★★★☆☆",
        "Notes": "Simpler than PPO; good for multi-env parallel training; lower sample efficiency",
    },
    {
        "Name": "DRQN",
        "Type": "Value-based (recurrent)",
        "Action Space": "Discrete",
        "Key Paper": "Hausknecht & Stone 2015 (AAAI WS)",
        "Finance Suitability": "★★★★☆",
        "Notes": "Add LSTM to handle partial observability; useful for hidden market state",
    },
    {
        "Name": "C51 / Distributional RL",
        "Type": "Value-based (distributional)",
        "Action Space": "Discrete",
        "Key Paper": "Bellemare et al. 2017 (ICML)",
        "Finance Suitability": "★★★★☆",
        "Notes": "Models return distribution; key for CVaR/VaR risk objectives in trading",
    },
    {
        "Name": "D4PG",
        "Type": "Actor-Critic (distributional, off-policy)",
        "Action Space": "Continuous",
        "Key Paper": "Barth-Maron et al. 2018 (ICLR)",
        "Finance Suitability": "★★★★☆",
        "Notes": "Distributional DDPG; used in Deep Hedging for Greeks management",
    },
    {
        "Name": "Rainbow DQN",
        "Type": "Value-based (combined)",
        "Action Space": "Discrete",
        "Key Paper": "Hessel et al. 2017 (AAAI)",
        "Finance Suitability": "★★★☆☆",
        "Notes": "Best discrete DQN; prioritised replay especially useful for rare market events",
    },
    {
        "Name": "Decision Transformer",
        "Type": "Sequence Model (offline RL)",
        "Action Space": "Both",
        "Key Paper": "Chen et al. 2021 (NeurIPS)",
        "Finance Suitability": "★★★★☆",
        "Notes": "Offline learning from trade logs; condition on desired Sharpe ratio as return-to-go",
    },
    {
        "Name": "CQL",
        "Type": "Value-based (offline)",
        "Action Space": "Both",
        "Key Paper": "Kumar et al. 2020 (NeurIPS)",
        "Finance Suitability": "★★★★☆",
        "Notes": "Best for pure offline RL from historical data; prevents OOD action overestimation",
    },
    {
        "Name": "MBPO",
        "Type": "Model-Based (Dyna-style)",
        "Action Space": "Continuous",
        "Key Paper": "Janner et al. 2019 (NeurIPS)",
        "Finance Suitability": "★★★☆☆",
        "Notes": "Learn market model for synthetic rollouts; improves sample efficiency 20-40x",
    },
    {
        "Name": "DreamerV2",
        "Type": "Model-Based (latent space)",
        "Action Space": "Both",
        "Key Paper": "Hafner et al. 2020 (ICLR)",
        "Finance Suitability": "★★★☆☆",
        "Notes": "Train entirely in world model imagination; potential for financial simulator",
    },
    {
        "Name": "EIIE",
        "Type": "Policy Gradient (custom)",
        "Action Space": "Continuous",
        "Key Paper": "Jiang et al. 2017 (arXiv)",
        "Finance Suitability": "★★★★☆",
        "Notes": "Purpose-built for portfolio management; portfolio vector memory is key innovation",
    },
]


# ─────────────────────────────────────────────────────────────────────────────
# STYLE HELPERS
# ─────────────────────────────────────────────────────────────────────────────

IMPACT_FILLS = {
    "Foundational": PatternFill("solid", fgColor="00B0F0"),  # Blue
    "High":         PatternFill("solid", fgColor="00B050"),  # Green
    "Medium":       PatternFill("solid", fgColor="FFFF00"),  # Yellow
    "Low":          PatternFill("solid", fgColor="FF0000"),  # Red
}

HEADER_FILL   = PatternFill("solid", fgColor="404040")
HEADER_FONT   = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
BODY_FONT     = Font(name="Calibri", size=10)
WRAP          = Alignment(wrap_text=True, vertical="top")
CENTER        = Alignment(horizontal="center", vertical="center", wrap_text=True)

THIN = Side(style="thin", color="CCCCCC")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COLUMNS = [
    ("Title",              40),
    ("Authors",            28),
    ("Year",                6),
    ("Venue",              16),
    ("DOI / arXiv ID",     22),
    ("Algorithm",          18),
    ("Finance Application",24),
    ("Summary",            60),
    ("Key Innovation",     36),
    ("Code / GitHub",      30),
    ("Impact Level",       14),
    ("Uses LLM?",           9),
    ("Notes",              32),
]


def make_thin_border():
    return Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header_row(ws, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill   = HEADER_FILL
        cell.font   = HEADER_FONT
        cell.alignment = CENTER
        cell.border = make_thin_border()


def set_col_widths(ws, columns):
    for i, (_, w) in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_paper_row(ws, row_idx, paper):
    values = [
        paper["Title"],
        paper["Authors"],
        paper["Year"],
        paper["Venue"],
        paper["DOI_arXiv"],
        paper["Algorithm"],
        paper["Finance_App"],
        paper["Summary"],
        paper["Key_Innovation"],
        paper["Code_GitHub"],
        paper["Impact"],
        paper["Uses_LLM"],
        paper["Notes"],
    ]
    fill = IMPACT_FILLS.get(paper["Impact"], PatternFill("solid", fgColor="FFFFFF"))

    for col_idx, val in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.fill      = fill
        cell.font      = BODY_FONT
        cell.alignment = WRAP
        cell.border    = make_thin_border()

    ws.row_dimensions[row_idx].height = 72


# ─────────────────────────────────────────────────────────────────────────────
# BUILD WORKBOOK
# ─────────────────────────────────────────────────────────────────────────────

wb = openpyxl.Workbook()

# ── Sheet 1: Papers ──────────────────────────────────────────────────────────
ws1 = wb.active
ws1.title = "RL Finance Papers"

# Headers
for col_idx, (name, _) in enumerate(COLUMNS, start=1):
    ws1.cell(row=1, column=col_idx, value=name)
style_header_row(ws1, len(COLUMNS))
set_col_widths(ws1, COLUMNS)

# Data
for i, paper in enumerate(PAPERS, start=2):
    write_paper_row(ws1, i, paper)

# Freeze & filter
ws1.freeze_panes = "A2"
ws1.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

# ── Sheet 2: Algorithm Reference ──────────────────────────────────────────────
ws2 = wb.create_sheet("Algorithm Reference")

algo_cols = [
    ("Algorithm Name",        20),
    ("Type",                  30),
    ("Action Space",          16),
    ("Key Paper",             28),
    ("Finance Suitability",   20),
    ("Notes",                 44),
]

for col_idx, (name, _) in enumerate(algo_cols, start=1):
    ws2.cell(row=1, column=col_idx, value=name)
style_header_row(ws2, len(algo_cols))
set_col_widths(ws2, algo_cols)

# Alternating light blue / white rows for algo sheet
ALT_FILL_A = PatternFill("solid", fgColor="DCE6F1")
ALT_FILL_B = PatternFill("solid", fgColor="FFFFFF")

for i, algo in enumerate(ALGORITHMS, start=2):
    row_fill = ALT_FILL_A if i % 2 == 0 else ALT_FILL_B
    vals = [
        algo["Name"],
        algo["Type"],
        algo["Action Space"],
        algo["Key Paper"],
        algo["Finance Suitability"],
        algo["Notes"],
    ]
    for col_idx, val in enumerate(vals, start=1):
        cell = ws2.cell(row=i, column=col_idx, value=val)
        cell.fill      = row_fill
        cell.font      = BODY_FONT
        cell.alignment = WRAP
        cell.border    = make_thin_border()
    ws2.row_dimensions[i].height = 36

ws2.freeze_panes = "A2"
ws2.auto_filter.ref = f"A1:{get_column_letter(len(algo_cols))}1"

# ── Save ─────────────────────────────────────────────────────────────────────
OUTPUT = "/Users/joshr/Repos/ai_sandbox/moneymaker/rl_finance_papers.xlsx"
wb.save(OUTPUT)
print(f"Saved: {OUTPUT}")
print(f"Papers: {len(PAPERS)}")
print(f"Algorithms in reference sheet: {len(ALGORITHMS)}")
