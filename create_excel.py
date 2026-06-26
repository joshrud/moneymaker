"""
Create algo_trading_repos.xlsx with data on top algorithmic trading GitHub repositories.
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# --- Data ---
# Fields: repo_name, github_url, stars, brokerage_api, uses_llm, sentiment_analysis,
#         summary, key_features, asset_types, backtesting, paper_trading

repos = [
    # Tier 1: 10k+ stars or very well known/maintained
    {
        "name": "freqtrade/freqtrade",
        "url": "https://github.com/freqtrade/freqtrade",
        "stars": 51800,
        "brokerage_api": "Binance, Bybit, Kraken, OKX, Gate.io, Bitget, HTX, Hyperliquid, BingX, Bitmart, CCXT",
        "uses_llm": "No",
        "sentiment": "No",
        "summary": "Free, open-source crypto trading bot written in Python with a rich feature set including FreqAI for adaptive machine learning strategy optimization, WebUI, and Telegram control.",
        "key_features": "FreqAI ML optimization, 300+ exchange support via CCXT, Hyperopt parameter tuning, WebUI, Telegram, backtesting, dry-run mode",
        "asset_types": "Crypto (spot & futures)",
        "backtesting": "Yes",
        "paper_trading": "Yes (dry-run)",
    },
    {
        "name": "microsoft/qlib",
        "url": "https://github.com/microsoft/qlib",
        "stars": 45200,
        "brokerage_api": "Yahoo Finance, Binance (data only)",
        "uses_llm": "Yes (RD-Agent LLM for factor mining)",
        "sentiment": "No",
        "summary": "Microsoft's AI-oriented quantitative investment platform covering the full ML pipeline: data processing, model training, backtesting, portfolio optimization, and order execution with 20+ built-in ML models.",
        "key_features": "20+ ML models (LSTM, Transformer, LightGBM, TabNet), RD-Agent LLM integration, alpha seeking, risk modeling, portfolio optimization, China & US markets",
        "asset_types": "Stocks (US & China)",
        "backtesting": "Yes",
        "paper_trading": "No",
    },
    {
        "name": "vnpy/vnpy",
        "url": "https://github.com/vnpy/vnpy",
        "stars": 42100,
        "brokerage_api": "CTP, XTP, Esunny, Interactive Brokers, Direct Futures, 15+ Chinese brokers",
        "uses_llm": "No",
        "sentiment": "No",
        "summary": "Python-based open-source quantitative trading system development framework with event-driven architecture, broad Chinese brokerage support, and vnpy.alpha ML suite for multi-factor strategies.",
        "key_features": "Event-driven architecture, CTA strategy engine, spread trading, options pricing/Greeks, TWAP/Iceberg/Sniper algos, ML alpha module (Lasso, LightGBM, MLP)",
        "asset_types": "Stocks, Futures, Options, ETFs, Crypto, Forex",
        "backtesting": "Yes",
        "paper_trading": "Yes",
    },
    {
        "name": "wilsonfreitas/awesome-quant",
        "url": "https://github.com/wilsonfreitas/awesome-quant",
        "stars": 27000,
        "brokerage_api": "N/A (curated list)",
        "uses_llm": "N/A",
        "sentiment": "N/A",
        "summary": "Curated list of 600+ insanely awesome libraries, packages, and resources for quantitative finance spanning data, backtesting, ML, derivatives pricing, and more.",
        "key_features": "600+ resources, covers Python/R/Julia/C++, data providers, backtesting, ML, derivatives, risk, portfolio optimization",
        "asset_types": "All",
        "backtesting": "N/A",
        "paper_trading": "N/A",
    },
    {
        "name": "nautechsystems/nautilus_trader",
        "url": "https://github.com/nautechsystems/nautilus_trader",
        "stars": 24200,
        "brokerage_api": "Binance, Coinbase, Kraken, Bybit, OKX, Interactive Brokers, dYdX, Hyperliquid, Betfair, Databento",
        "uses_llm": "No",
        "sentiment": "No",
        "summary": "Production-grade, Rust-native trading engine with Python control plane. Deterministic event-driven architecture, nanosecond-resolution backtesting, and identical strategy code for research and live deployment.",
        "key_features": "Rust core (ultra-fast), Python API, 25+ exchange adapters, asset-class-agnostic, nanosecond backtesting, multi-venue strategies, Redis persistence, Docker support",
        "asset_types": "Crypto, Stocks, Forex, Futures, Options, Sports Betting",
        "backtesting": "Yes",
        "paper_trading": "Yes",
    },
    {
        "name": "mementum/backtrader",
        "url": "https://github.com/mementum/backtrader",
        "stars": 22100,
        "brokerage_api": "Interactive Brokers, OANDA, Visual Chart",
        "uses_llm": "No",
        "sentiment": "No",
        "summary": "Feature-rich Python backtesting library with 122 built-in indicators, multiple timeframes, and broker simulation. Widely used and well-documented but no longer actively developed.",
        "key_features": "122+ indicators, multiple data feeds, broker simulation, slippage modeling, OCO/stop/trail orders, matplotlib plotting, Analyzers API, Strategy Optimization",
        "asset_types": "Stocks, Forex, Futures, Crypto",
        "backtesting": "Yes",
        "paper_trading": "No",
    },
    {
        "name": "QuantConnect/Lean",
        "url": "https://github.com/QuantConnect/Lean",
        "stars": 20200,
        "brokerage_api": "Alpaca, Interactive Brokers, TD Ameritrade/Schwab, Coinbase, Binance, Kraken, Oanda, Bitfinex, Tradier",
        "uses_llm": "No",
        "sentiment": "No",
        "summary": "Professional-grade algorithmic trading engine by QuantConnect supporting Python and C#. Powers QuantConnect's cloud platform with extensive brokerage integrations and alternative data support.",
        "key_features": "Multi-language (Python/C#), 10+ brokerage integrations, options/futures/forex, alternative data, Jupyter research, CLI, cloud-hybrid deployment",
        "asset_types": "Stocks, Options, Futures, Forex, Crypto, CFDs",
        "backtesting": "Yes",
        "paper_trading": "Yes",
    },
    {
        "name": "AI4Finance-Foundation/FinGPT",
        "url": "https://github.com/AI4Finance-Foundation/FinGPT",
        "stars": 20700,
        "brokerage_api": "N/A (research/LLM framework)",
        "uses_llm": "Yes (core product - financial LLM)",
        "sentiment": "Yes (financial sentiment classification)",
        "summary": "Open-source financial large language model framework with instruction tuning, sentiment analysis, stock movement prediction, and RAG for financial news and SEC filings.",
        "key_features": "FinBERT/LLaMA/Qwen fine-tuning, LoRA on consumer GPUs, sentiment classification, stock forecasting, RAG, multi-task NLP, <$300 fine-tuning cost",
        "asset_types": "Stocks (US)",
        "backtesting": "No",
        "paper_trading": "No",
    },
    {
        "name": "stefan-jansen/machine-learning-for-trading",
        "url": "https://github.com/stefan-jansen/machine-learning-for-trading",
        "stars": 19300,
        "brokerage_api": "Interactive Brokers, Alpaca, QuantConnect",
        "uses_llm": "Yes (FinBERT, RAG, multi-agent)",
        "sentiment": "Yes (FinBERT, NLP financial text)",
        "summary": "Comprehensive ML-to-production trading workflow from the book 'Machine Learning for Algorithmic Trading'. Covers data infrastructure, feature engineering, model training, strategy implementation, and live deployment.",
        "key_features": "19+ data providers, FinBERT sentiment, XGBoost/LightGBM/LSTM/Transformer models, SEC filing RAG, 9 case studies, walk-forward validation, production deployment",
        "asset_types": "Stocks, ETFs, Crypto, Options, Forex, Futures",
        "backtesting": "Yes",
        "paper_trading": "No",
    },
    {
        "name": "hummingbot/hummingbot",
        "url": "https://github.com/hummingbot/hummingbot",
        "stars": 19000,
        "brokerage_api": "Binance, Coinbase, Bybit, KuCoin, OKX, Gate.io, dYdX, Hyperliquid, Uniswap, Curve, Raydium, 50+ exchanges",
        "uses_llm": "No (MCP integration for Claude/Gemini)",
        "sentiment": "No",
        "summary": "Open-source framework for building and deploying high-frequency trading strategies across 50+ CEX and DEX venues. Generated over $34B in trading volume with support for market making, arbitrage, and AMM strategies.",
        "key_features": "50+ exchange connectors, CEX & DEX, market making, arbitrage, AMM liquidity, HFT, paper trading, $34B+ volume processed, MCP server for AI assistants",
        "asset_types": "Crypto (spot, perpetuals, DEX)",
        "backtesting": "Yes",
        "paper_trading": "Yes",
    },
    {
        "name": "AI4Finance-Foundation/FinRL",
        "url": "https://github.com/AI4Finance-Foundation/FinRL",
        "stars": 15500,
        "brokerage_api": "Alpaca",
        "uses_llm": "No",
        "sentiment": "No",
        "summary": "First open-source deep reinforcement learning framework for financial trading. Implements A2C, DDPG, PPO, SAC, TD3 agents with a train-test-trade pipeline and 14+ technical indicators.",
        "key_features": "DRL algorithms (A2C, DDPG, PPO, SAC, TD3), 14+ technical indicators, multi-asset portfolio, ensemble strategies, 14+ data providers, backtesting",
        "asset_types": "Stocks, Crypto",
        "backtesting": "Yes",
        "paper_trading": "No",
    },
    {
        "name": "goldmansachs/gs-quant",
        "url": "https://github.com/goldmansachs/gs-quant",
        "stars": 10900,
        "brokerage_api": "Goldman Sachs internal API (institutional access required)",
        "uses_llm": "No",
        "sentiment": "No",
        "summary": "Goldman Sachs Python toolkit for quantitative finance used by 1000+ internal quant developers. Specializes in derivatives structuring, trading, and risk management built on 25 years of GS market experience.",
        "key_features": "Derivatives pricing, risk management, scenario analysis, backtesting, portfolio analytics, GS Marquee platform integration, institutional-grade",
        "asset_types": "Derivatives, Equities, Fixed Income, FX, Commodities",
        "backtesting": "Yes",
        "paper_trading": "No",
    },
    # Tier 2: 2k-10k stars
    {
        "name": "jesse-ai/jesse",
        "url": "https://github.com/jesse-ai/jesse",
        "stars": 8100,
        "brokerage_api": "Binance, Bybit, FTX (historical), Bitget",
        "uses_llm": "Yes (JesseGPT AI assistant)",
        "sentiment": "No",
        "summary": "Advanced crypto trading framework with 300+ technical indicators, Monte Carlo simulation for strategy stress testing, ML pipeline, and JesseGPT AI assistant for writing and optimizing strategies.",
        "key_features": "300+ indicators, Monte Carlo simulation, ML pipeline, JesseGPT AI, partial fills, leveraged/short trading, strategy debugging, paper trading",
        "asset_types": "Crypto (spot & futures)",
        "backtesting": "Yes",
        "paper_trading": "Yes",
    },
    {
        "name": "sammchardy/python-binance",
        "url": "https://github.com/sammchardy/python-binance",
        "stars": 7200,
        "brokerage_api": "Binance (spot, futures, margin, options)",
        "uses_llm": "No",
        "sentiment": "No",
        "summary": "Comprehensive Python client for the Binance API covering all REST and WebSocket endpoints for spot, margin, futures, and options trading with asyncio support.",
        "key_features": "Full Binance REST + WebSocket API, asyncio, RSA/EDDSA auth, testnet support, Depth Cache, historical klines, margin/futures/options, proxy support",
        "asset_types": "Crypto (Binance only)",
        "backtesting": "No",
        "paper_trading": "No (testnet support)",
    },
    {
        "name": "tensortrade-org/tensortrade",
        "url": "https://github.com/tensortrade-org/tensortrade",
        "stars": 6400,
        "brokerage_api": "N/A (RL research framework)",
        "uses_llm": "No",
        "sentiment": "No",
        "summary": "Reinforcement learning framework for training, evaluating, and deploying trading agents. Composable components (environments, action schemes, reward functions) with support for PPO via Ray RLlib.",
        "key_features": "RL training (PPO, DQN), composable environment components, Ray RLlib distributed training, Optuna hyperparameter optimization, walk-forward validation",
        "asset_types": "Crypto (primarily BTC/USD)",
        "backtesting": "Yes",
        "paper_trading": "No",
    },
    {
        "name": "Drakkar-Software/OctoBot",
        "url": "https://github.com/Drakkar-Software/OctoBot",
        "stars": 6100,
        "brokerage_api": "Binance, Coinbase, KuCoin, Bybit, OKX, MEXC, Hyperliquid, 15+ exchanges",
        "uses_llm": "Yes (ChatGPT / Ollama integration)",
        "sentiment": "Yes (Google Trends, Reddit signals)",
        "summary": "Open-source crypto trading bot with AI strategy support via ChatGPT/Ollama, grid/DCA/basket strategies, TradingView integration, and social sentiment signals from Google Trends and Reddit.",
        "key_features": "ChatGPT/Ollama LLM strategies, grid/DCA trading, TradingView alerts, Google Trends/Reddit sentiment, 15+ exchanges, paper trading, cloud + local deployment",
        "asset_types": "Crypto",
        "backtesting": "Yes",
        "paper_trading": "Yes",
    },
    {
        "name": "Superalgos/Superalgos",
        "url": "https://github.com/Superalgos/Superalgos",
        "stars": 6100,
        "brokerage_api": "Binance, multiple crypto exchanges",
        "uses_llm": "No",
        "sentiment": "No",
        "summary": "Free open-source crypto trading platform with visual strategy designer, integrated charting system, data mining, backtesting, paper trading, and multi-server deployment.",
        "key_features": "Visual strategy designer, integrated charts, data mining, multi-server farm deployment, Node.js-based, no code/low code approach",
        "asset_types": "Crypto",
        "backtesting": "Yes",
        "paper_trading": "Yes",
    },
    {
        "name": "ccxt/ccxt",
        "url": "https://github.com/ccxt/ccxt",
        "stars": 42000,
        "brokerage_api": "100+ crypto exchanges (Binance, Coinbase, Kraken, Bybit, OKX, KuCoin, HTX, Gate.io, etc.)",
        "uses_llm": "No",
        "sentiment": "No",
        "summary": "The de-facto standard unified cryptocurrency exchange trading API library supporting 100+ exchanges in Python, JavaScript, PHP, Go, C#, and Java. Not a trading bot but essential infrastructure.",
        "key_features": "100+ exchange connectors, unified REST + WebSocket API, spot/futures/margin, multi-language, CCXT Pro for streaming, orjson for performance",
        "asset_types": "Crypto (all asset types on supported exchanges)",
        "backtesting": "No",
        "paper_trading": "No",
    },
    {
        "name": "gbeced/pyalgotrade",
        "url": "https://github.com/gbeced/pyalgotrade",
        "stars": 4700,
        "brokerage_api": "Bitstamp (deprecated/archived)",
        "uses_llm": "No",
        "sentiment": "No",
        "summary": "ARCHIVED: Python algorithmic trading library. Deprecated in November 2023. Was a solid event-driven backtesting framework with TA-Lib integration and Twitter event handling.",
        "key_features": "Event-driven backtesting, SMA/EMA/RSI/Bollinger Bands, TA-Lib integration, Sharpe/drawdown analysis, Twitter events (historical)",
        "asset_types": "Stocks, Crypto (Bitstamp)",
        "backtesting": "Yes",
        "paper_trading": "Yes",
    },
    {
        "name": "alpacahq/alpaca-py",
        "url": "https://github.com/alpacahq/alpaca-py",
        "stars": 1400,
        "brokerage_api": "Alpaca (stocks, crypto, options)",
        "uses_llm": "No",
        "sentiment": "No",
        "summary": "Official Python SDK for Alpaca's commission-free trading API. Supports stocks, crypto, and options with async/await, Pydantic validation, paper trading, and WebSocket streaming.",
        "key_features": "Async/await, Pydantic validation, paper trading, WebSocket streaming, 5000+ stocks, 20+ crypto, options, Broker API for fintech apps",
        "asset_types": "Stocks, Crypto, Options",
        "backtesting": "No",
        "paper_trading": "Yes",
    },
    {
        "name": "Lumiwealth/lumibot",
        "url": "https://github.com/Lumiwealth/lumibot",
        "stars": 1700,
        "brokerage_api": "Alpaca, Interactive Brokers, Tradier, Schwab, Tradovate, TopstepX, Coinbase, Kraken, Binance, Bybit, KuCoin",
        "uses_llm": "Yes (built-in AI agent runtime)",
        "sentiment": "No",
        "summary": "Backtestable AI trading agent framework for stocks, options, crypto, futures, and forex. Same code runs for backtesting and live trading with built-in LLM agent runtime supporting multiple providers.",
        "key_features": "LLM agent runtime, same code backtest+live, SEC filings, FRED macro data, DuckDB analytics, multi-broker, Telegram notifications, options/futures support",
        "asset_types": "Stocks, Options, Crypto, Futures, Forex",
        "backtesting": "Yes",
        "paper_trading": "Yes",
    },
    {
        "name": "blankly-finance/blankly",
        "url": "https://github.com/blankly-finance/blankly",
        "stars": 2500,
        "brokerage_api": "Coinbase Pro, Binance, Alpaca, OANDA, KuCoin, OKX, Kraken",
        "uses_llm": "No",
        "sentiment": "No",
        "summary": "Unified Python framework for building, backtesting, and deploying trading algorithms across multiple exchanges with identical code for all modes (live, paper, sandbox, backtest).",
        "key_features": "Single-line mode switching, multi-exchange, event-driven, pre-built RSI/MACD/Golden Cross strategies, CLI deployment, WebSocket support",
        "asset_types": "Stocks, Crypto, Forex, Futures",
        "backtesting": "Yes",
        "paper_trading": "Yes",
    },
    {
        "name": "stefan-jansen/zipline-reloaded",
        "url": "https://github.com/stefan-jansen/zipline-reloaded",
        "stars": 1700,
        "brokerage_api": "None (backtesting only)",
        "uses_llm": "No",
        "sentiment": "No",
        "summary": "Maintained fork of Quantopian's Zipline backtesting library by the author of 'ML for Trading'. Event-driven system deeply integrated with pandas/PyData ecosystem.",
        "key_features": "Event-driven backtesting, pandas integration, PyData ecosystem, Python 3.9+, Quantopian-compatible strategies, Pipeline API, custom data bundles",
        "asset_types": "Stocks, ETFs",
        "backtesting": "Yes",
        "paper_trading": "No",
    },
    {
        "name": "quantopian/zipline",
        "url": "https://github.com/quantopian/zipline",
        "stars": 17500,
        "brokerage_api": "None (archived, backtesting only)",
        "uses_llm": "No",
        "sentiment": "No",
        "summary": "ARCHIVED: Original Quantopian backtesting library. Event-driven, Pythonic algorithmic trading backtester. No longer maintained since Quantopian shut down in 2020.",
        "key_features": "Event-driven backtesting, Pipeline API, pandas integration, factor research, Quantopian ecosystem (archived)",
        "asset_types": "Stocks, ETFs",
        "backtesting": "Yes",
        "paper_trading": "No",
    },
    {
        "name": "AI4Finance-Foundation/FinRL-Trading",
        "url": "https://github.com/AI4Finance-Foundation/FinRL-Trading",
        "stars": 800,
        "brokerage_api": "Alpaca",
        "uses_llm": "No",
        "sentiment": "No",
        "summary": "AI-native modular infrastructure for quantitative trading (FinRL-X). Production-focused successor to FinRL with modular design for deploying DRL trading agents.",
        "key_features": "Modular DRL infrastructure, production deployment, Alpaca integration, multi-agent ensemble, AI-native design",
        "asset_types": "Stocks",
        "backtesting": "Yes",
        "paper_trading": "Yes",
    },
    {
        "name": "google/tf-quant-finance",
        "url": "https://github.com/google/tf-quant-finance",
        "stars": 4500,
        "brokerage_api": "None (quantitative finance library)",
        "uses_llm": "No",
        "sentiment": "No",
        "summary": "Google's TensorFlow-based quantitative finance library for high-performance derivatives pricing, Monte Carlo simulation, and interest rate model calibration on GPU/TPU.",
        "key_features": "TensorFlow GPU/TPU acceleration, Monte Carlo pricing, Black-Scholes, HJM/Hull-White rates models, American options, batched computations",
        "asset_types": "Derivatives, Options, Fixed Income",
        "backtesting": "No",
        "paper_trading": "No",
    },
    {
        "name": "hftbacktest/hftbacktest",
        "url": "https://github.com/nkaz001/hftbacktest",
        "stars": 4100,
        "brokerage_api": "Binance, Bybit (data feeds)",
        "uses_llm": "No",
        "sentiment": "No",
        "summary": "High-frequency trading and market making backtester with nanosecond-resolution, L2 order book simulation, and Rust core for maximum performance.",
        "key_features": "Nanosecond resolution, L2 order book, latency simulation, Rust backend, Python interface, HFT-grade accuracy, market microstructure modeling",
        "asset_types": "Crypto (HFT focus)",
        "backtesting": "Yes",
        "paper_trading": "No",
    },
    {
        "name": "alpacahq/alpaca-trade-api-python",
        "url": "https://github.com/alpacahq/alpaca-trade-api-python",
        "stars": 1800,
        "brokerage_api": "Alpaca (legacy SDK)",
        "uses_llm": "No",
        "sentiment": "No",
        "summary": "Legacy Python SDK for Alpaca's commission-free trading API. Superseded by alpaca-py but still widely used. Supports REST and WebSocket streaming for stocks and crypto.",
        "key_features": "REST + WebSocket, paper trading, polygon data integration (legacy), streaming market data, order management",
        "asset_types": "Stocks, Crypto",
        "backtesting": "No",
        "paper_trading": "Yes",
    },
]


def get_row_color(stars, name):
    """Return fill color based on star count and special cases."""
    # Special cases for well-known/maintained regardless of stars
    blue_always = {
        "freqtrade/freqtrade", "microsoft/qlib", "vnpy/vnpy",
        "nautechsystems/nautilus_trader", "mementum/backtrader",
        "QuantConnect/Lean", "AI4Finance-Foundation/FinGPT",
        "stefan-jansen/machine-learning-for-trading", "hummingbot/hummingbot",
        "AI4Finance-Foundation/FinRL", "goldmansachs/gs-quant",
        "quantopian/zipline", "ccxt/ccxt",
        "wilsonfreitas/awesome-quant",
    }
    if name in blue_always or stars >= 10000:
        return "00B0F0"  # Blue
    elif 2000 <= stars < 10000:
        return "00B050"  # Green
    elif 500 <= stars < 2000:
        return "FFFF00"  # Yellow
    else:
        return "FF0000"  # Red


def make_fill(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


def create_workbook():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Algo Trading Repos"

    # --- Column headers ---
    headers = [
        "Repo Name", "GitHub URL", "Stars (approx)", "Brokerage / Data API",
        "Uses LLM?", "Sentiment Analysis?", "Summary", "Key Features",
        "Asset Types", "Backtesting?", "Paper Trading?",
    ]

    # Header style
    header_fill = PatternFill(start_color="404040", end_color="404040", fill_type="solid")
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="top", wrap_text=True)

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

    ws.row_dimensions[1].height = 30

    # --- Data rows ---
    for row_idx, repo in enumerate(repos, start=2):
        color = get_row_color(repo["stars"], repo["name"])
        row_fill = make_fill(color)
        # Darker font for yellow rows for readability
        if color == "FFFF00":
            data_font = Font(name="Calibri", size=10, color="000000")
        elif color in ("FF0000",):
            data_font = Font(name="Calibri", size=10, color="000000")
        else:
            data_font = Font(name="Calibri", size=10, color="000000")

        values = [
            repo["name"],
            repo["url"],
            repo["stars"],
            repo["brokerage_api"],
            repo["uses_llm"],
            repo["sentiment"],
            repo["summary"],
            repo["key_features"],
            repo["asset_types"],
            repo["backtesting"],
            repo["paper_trading"],
        ]

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = row_fill
            cell.font = data_font
            cell.border = thin_border

            # Alignment
            if col_idx in (7, 8):  # Summary and Key Features - wrap
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            elif col_idx in (1, 4, 9):
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            # Hyperlink on URL column
            if col_idx == 2:
                cell.hyperlink = value
                cell.style = "Hyperlink"
                cell.fill = row_fill  # re-apply after hyperlink style reset
                cell.font = Font(name="Calibri", size=10, color="0563C1", underline="single")

        ws.row_dimensions[row_idx].height = 55

    # --- Column widths ---
    col_widths = {
        1: 35,   # Repo Name
        2: 50,   # GitHub URL
        3: 14,   # Stars
        4: 45,   # Brokerage
        5: 12,   # LLM
        6: 16,   # Sentiment
        7: 60,   # Summary
        8: 60,   # Key Features
        9: 25,   # Asset Types
        10: 12,  # Backtesting
        11: 14,  # Paper Trading
    }
    for col, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # --- Add a legend sheet ---
    legend_ws = wb.create_sheet("Legend")
    legend_ws.column_dimensions["A"].width = 30
    legend_ws.column_dimensions["B"].width = 50

    legend_data = [
        ("Color", "Meaning"),
        ("Blue (00B0F0)", "Best repos: 10,000+ stars OR very well known/actively maintained"),
        ("Green (00B050)", "Decent repos: 2,000–10,000 stars"),
        ("Yellow (FFFF00)", "Mediocre: 500–2,000 stars"),
        ("Red (FF0000)", "Lowest: under 500 stars or abandoned/archived"),
    ]

    legend_header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    for row_idx, (col_a, col_b) in enumerate(legend_data, start=1):
        ca = legend_ws.cell(row=row_idx, column=1, value=col_a)
        cb = legend_ws.cell(row=row_idx, column=2, value=col_b)
        if row_idx == 1:
            for c in (ca, cb):
                c.fill = make_fill("404040")
                c.font = legend_header_font
                c.alignment = center_align
        else:
            color_map = {
                "Blue (00B0F0)": "00B0F0",
                "Green (00B050)": "00B050",
                "Yellow (FFFF00)": "FFFF00",
                "Red (FF0000)": "FF0000",
            }
            hex_c = color_map.get(col_a, "FFFFFF")
            fill = make_fill(hex_c)
            ca.fill = fill
            ca.font = Font(name="Calibri", size=10)
            cb.fill = fill
            cb.font = Font(name="Calibri", size=10)
        for c in (ca, cb):
            c.border = thin_border
        legend_ws.row_dimensions[row_idx].height = 20

    output_path = "/Users/joshr/Repos/ai_sandbox/moneymaker/algo_trading_repos.xlsx"
    wb.save(output_path)
    print(f"Excel file saved to: {output_path}")
    print(f"Total repos: {len(repos)}")


if __name__ == "__main__":
    create_workbook()
